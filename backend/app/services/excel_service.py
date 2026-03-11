from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.models.user import Finance


def build_finance_excel(finances: Iterable[Finance]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Finances"

    headers = ["Date", "Type", "Description", "Category", "Amount"]
    ws.append(headers)

    for finance in finances:
        ws.append(
            [
                finance.entry_date.strftime("%Y-%m-%d"),
                finance.entry_type.title(),
                finance.description,
                finance.category,
                float(finance.amount),
            ]
        )

    for col_idx, header in enumerate(headers, start=1):
        column = get_column_letter(col_idx)
        ws[f"{column}1"].font = ws[f"{column}1"].font.copy(bold=True)
        ws.column_dimensions[column].width = max(len(header) + 2, 15)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
